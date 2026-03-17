import json
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


def get_output_filename(url: str) -> str:
    """Извлекает последний сегмент из URL и заменяет '-' на '_'"""
    parsed = urlparse(url)
    path = parsed.path.rstrip('/')
    last_segment = path.split('/')[-1] if path else 'vtb_data'
    if not last_segment:
        last_segment = parsed.netloc.replace('.', '_')
    return last_segment.replace('-', '_')


def parse_cards(page, base_url: str) -> list[dict]:
    """Парсит все типы карточек: рекомендации, сеточные блоки, chips-контент"""
    cards = []
    
    # === Селекторы ===
    # 1. Карточки-рекомендации (ссылки)
    rec_selector = 'a.cardstyles__Link-card-recommendation__sc-xkw79f-1'
    # 2. Универсальный селектор для контент-карточек (ловит и grid, и chips)
    # Ищем по внутреннему шаблону, который одинаков у всех типов
    card_template_selector = 'div.card-mediumstyles__CardTemplate-card-base__sc-senydt-0'
    
    # === Парсинг карточек-рекомендаций ===
    for el in page.query_selector_all(rec_selector):
        try:
            title_el = el.query_selector('[data-ym-card-recommendation-title="true"]')
            title = title_el.text_content().strip() if title_el else el.text_content().strip().split('\n')[0]
            
            href = el.get_attribute('data-ym-card-recommendation-href') or el.get_attribute('href')
            if href and not href.startswith(('http://', 'https://')):
                href = urljoin(base_url, href)
            
            if title and href:
                cards.append({
                    "description": title,
                    "website_name": "",
                    "Link": href.strip(),
                    "elements_list": [],
                    "FAQ": []
                })
        except Exception:
            continue
    
    # === Парсинг контент-карточек (универсальный подход) ===
    for card in page.query_selector_all(card_template_selector):
        try:
            # Пропускаем, если это уже обработанная рекомендация (вложенная в <a>)
            if card.query_selector('a.cardstyles__Link-card-recommendation__sc-xkw79f-1'):
                continue
                
            # Заголовок
            title_el = card.query_selector('p.typographystyles__Box-foundation-kit__sc-14qzghz-0')
            if not title_el:
                continue
            title = title_el.text_content().strip()
            if not title:
                continue
            
            # Описание (пробуем несколько вариантов вложенности)
            description = ""
            desc_selectors = [
                'div.markdownstyles__StyledReactMarkdown-foundation-kit__sc-v45gkz-0 div.markdown-paragraphstyles__ParagraphTypography-foundation-kit__sc-otngat-0',
                'div.markdownstyles__StyledReactMarkdown-foundation-kit__sc-v45gkz-0 div.typographystyles__Box-foundation-kit__sc-14qzghz-0',
                'div.markdownstyles__StyledReactMarkdown-foundation-kit__sc-v45gkz-0',
            ]
            for sel in desc_selectors:
                desc_el = card.query_selector(sel)
                if desc_el:
                    description = desc_el.text_content().strip()
                    break
            
            # Ссылки из кнопок
            links = []
            for btn in card.query_selector_all('a.buttonstyles__LinkBox-foundation-kit__sc-sa2uer-1'):
                href = btn.get_attribute('href')
                if href:
                    if not href.startswith(('http://', 'https://')):
                        href = urljoin(base_url, href)
                    btn_text = btn.text_content().strip().lower()
                    links.append((href, btn_text))
            
            # Приоритет ссылок: 1) "Подробнее", 2) первая информационная, 3) любая
            main_link = ""
            for href, text in links:
                if "подробнее" in text or "узнать больше" in text:
                    main_link = href
                    break
            if not main_link and links:
                # Берём первую ссылку, которая не ведёт на online.vtb.ru (если есть выбор)
                for href, _ in links:
                    if "online.vtb.ru" not in href:
                        main_link = href
                        break
                if not main_link:
                    main_link = links[0][0]
            
            # Формируем описание
            full_description = f"{title}\n{description}".strip() if description else title
            
            cards.append({
                "description": full_description,
                "website_name": "",
                "Link": main_link.strip(),
                "elements_list": [],
                "FAQ": [],
                "_all_links": [l[0] for l in links]  # Опционально: все ссылки для отладки
            })
        except Exception:
            continue
    
    return cards


def parse_website_name(page) -> str:
    """
    Извлекает главный заголовок страницы из hero-секции.
    Возвращает текст из <h1> внутри <section id="hero">
    """
    try:
        # Приоритет: h1 с классом внутри hero-секции
        title_el = page.query_selector('section#hero h1.typographystyles__Box-foundation-kit__sc-14qzghz-0')
        if title_el:
            return title_el.text_content().strip()
        
        # Фолбэк: любой h1 в секции #hero
        title_el = page.query_selector('section#hero h1')
        if title_el:
            return title_el.text_content().strip()
        
        # Последний фолбэк: первый h1 на странице
        title_el = page.query_selector('h1')
        if title_el:
            return title_el.text_content().strip()
    except Exception:
        pass
    
    return ""


def parse_page_elements(page) -> list[str]:
    """Парсит элементы с описанием услуг на странице"""
    elements_data = []
    container_selector = 'div.card-mediumstyles__InnerGroupStyled-card-base__sc-senydt-2'
    
    containers = page.query_selector_all(container_selector)
    
    for container in containers:
        try:
            title_el = container.query_selector('p.typographystyles__Box-foundation-kit__sc-14qzghz-0')
            if not title_el:
                continue
            title = title_el.text_content().strip()
            if not title:
                continue
            
            desc_el = container.query_selector('div.markdownstyles__StyledReactMarkdown-foundation-kit__sc-v45gkz-0 div.markdown-paragraphstyles__ParagraphTypography-foundation-kit__sc-otngat-0')
            description = desc_el.text_content().strip() if desc_el else ""
            
            elements_data.append(f"{title}\n{description}")
        except Exception:
            continue
    
    # Фолбэк-селектор
    if not elements_data:
        blocks = page.query_selector_all('div.card-mediumstyles__ParentGroupStyled-card-base__sc-senydt-1')
        for block in blocks:
            try:
                title_el = block.query_selector('p.typographystyles__Box-foundation-kit__sc-14qzghz-0')
                if title_el:
                    title = title_el.text_content().strip()
                    desc_el = block.query_selector('div.markdownstyles__StyledReactMarkdown-foundation-kit__sc-v45gkz-0 div.typographystyles__Box-foundation-kit__sc-14qzghz-0')
                    description = desc_el.text_content().strip() if desc_el else ""
                    if title:
                        elements_data.append(f"{title}\n{description}")
            except Exception:
                continue
    
    return elements_data


def parse_faq_elements(page) -> list[str]:
    """Парсит FAQ из аккордеонов на странице, учитывая структуру с табами"""
    faq_data = []
    
    tab_content_selectors = [
        'div.tabs-layoutstyles__TabContentContainer-foundation-kit__sc-hj413w-2',
        'div.tabstyles__TabsContainer-foundation-kit__sc-1hmeyb5-0',
    ]
    
    all_accordions = []
    
    for tab_content_selector in tab_content_selectors:
        tab_containers = page.query_selector_all(tab_content_selector)
        for container in tab_containers:
            accordions = container.query_selector_all('div.accordionstyles__BoxOuter-accordion__sc-1d34irg-1')
            all_accordions.extend(accordions)
    
    if not all_accordions:
        all_accordions = page.query_selector_all('div.accordionstyles__BoxOuter-accordion__sc-1d34irg-1')
    
    for accordion in all_accordions:
        try:
            question_el = accordion.query_selector('h2[itemprop="name"]')
            if not question_el:
                continue
            question = question_el.text_content().strip()
            if not question:
                continue
            
            answer_el = accordion.query_selector('div[itemprop="text"]')
            if not answer_el:
                continue
            answer = answer_el.text_content().strip()
            answer = ' '.join(answer.split())
            
            if answer:
                faq_data.append(f"{question}\n{answer}")
        except Exception:
            continue
    
    # Удаляем дубликаты
    seen = set()
    unique_faq = []
    for item in faq_data:
        if item not in seen:
            seen.add(item)
            unique_faq.append(item)
    
    return unique_faq


def parse_detail_page(url: str, context, wait_time: float = 2.0) -> tuple[str, list[str], list[str]]:
    """
    Открывает страницу и парсит: website_name, элементы + FAQ.
    Возвращает кортеж: (website_name, elements_list, FAQ_list)
    """
    website_name, elements, faq = "", [], []
    
    try:
        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_selector("body", timeout=10000)
        
        if wait_time > 0:
            page.wait_for_timeout(int(wait_time * 1000))
        
        # Скролл для ленивой загрузки
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(1000)
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(500)
        
        # Парсим все три типа данных
        website_name = parse_website_name(page)
        elements = parse_page_elements(page)
        faq = parse_faq_elements(page)
        
        page.close()
        return website_name, elements, faq
        
    except PlaywrightTimeout:
        print(f"   ⏱ Таймаут при загрузке {url}")
        return "", [], []
    except Exception as e:
        print(f"   ❌ Ошибка при парсинге {url}: {e}")
        return "", [], []


def json_to_xlsx(json_file: str, output_file: str = None):
    """
    Конвертирует JSON с карточками в XLSX файл.
    
    Новая структура Excel (6 столбцов):
    | description | website_name | Link | list_type | title | content |
    """
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "VTB platez&perev"
    
    # === Заголовки (6 столбцов) ===
    headers = ["description", "website_name", "Link", "list_type", "title", "content"]
    ws.append(headers)
    
    # Стилизация заголовков
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # === Заполнение данными ===
    for card in data:
        description = card.get('description', '')
        website_name = card.get('website_name', '')  # <-- НОВОЕ ПОЛЕ
        link = card.get('Link', '')
        elements_list = card.get('elements_list', [])
        faq_list = card.get('FAQ', [])
        
        # Обрабатываем elements_list
        for item in elements_list:
            if '\n' in item:
                title, content = item.split('\n', 1)
            else:
                title = item
                content = ""
            
            row = [description, website_name, link, 'elements_list', title.strip(), content.strip()]
            ws.append(row)
        
        # Обрабатываем FAQ
        for item in faq_list:
            if '\n' in item:
                title, content = item.split('\n', 1)
            else:
                title = item
                content = ""
            
            row = [description, website_name, link, 'FAQ', title.strip(), content.strip()]
            ws.append(row)
    
    # === Перенос текста и границы ===
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
    
    # Сохранение промежуточного файла
    if output_file is None:
        output_file = Path(json_file).with_suffix('.xlsx')
    else:
        output_file = Path(output_file)
    
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    # === АВТОНАСТРОЙКА ШИРИНЫ СТОЛБЦОВ ===
    wb = load_workbook(output_file)
    ws = wb.active
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    
    print(f"💾 Excel сохранён: {output_file.resolve()}")
    print(f"📊 Всего строк данных: {ws.max_row - 1}")
    
    return output_file


def main(
    url: str,
    headless: bool = False,
    output_dir: str = None,
    wait_time: float = 2.0,
    detail_wait_time: float = 2.0,
    delay_between_requests: float = 1.5
):
    """Главная функция парсинга с сохранением JSON и XLSX в подпапку"""
    print(f"🚀 Парсинг: {url}")
    
    # Базовая директория для сохранения
    if output_dir is None:
        base_output_dir = Path(__file__).parent / "vtb_downloads"
    else:
        base_output_dir = Path(output_dir)
    base_output_dir.mkdir(parents=True, exist_ok=True)
    
    # Генерируем имя подпапки из URL
    folder_name = get_output_filename(url)
    
    # Создаём подпапку внутри base_output_dir
    output_folder = base_output_dir / folder_name
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # Пути к файлам внутри подпапки
    json_path = output_folder / f"{folder_name}.json"
    xlsx_path = output_folder / f"{folder_name}.xlsx"
    
    cards = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            viewport={"width": 1920, "height": 1080}
        )
        page = context.new_page()
        
        try:
            # === ШАГ 1: Парсинг главной страницы ===
            print("   📡 Загрузка главной страницы...")
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_selector("body", timeout=10000)
            
            if wait_time > 0:
                page.wait_for_timeout(int(wait_time * 1000))
            
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1000)
            page.evaluate("window.scrollTo(0, 0)")
            page.wait_for_timeout(500)
            
            print("   🔎 Поиск карточек...")
            cards = parse_cards(page, url)
            print(f"📈 Найдено карточек: {len(cards)}")
            
            # === ШАГ 2: Переход по ссылкам и парсинг деталей ===
            if cards:
                print(f"\n🔗 Переход по ссылкам для детального парсинга...")
                for i, card in enumerate(cards, 1):
                    print(f"   [{i}/{len(cards)}] {card['description']}")
                    
                    # Теперь получаем 3 значения: website_name, elements, faq
                    website_name, elements, faq = parse_detail_page(card['Link'], context, wait_time=detail_wait_time)
                    
                    card['website_name'] = website_name  # <-- Заполняем новое поле
                    card['elements_list'] = elements
                    card['FAQ'] = faq
                    
                    print(f"      ✓ Заголовок: {website_name[:50]}..., Элементов: {len(elements)}, FAQ: {len(faq)}")
                    
                    if i < len(cards) and delay_between_requests > 0:
                        time.sleep(delay_between_requests)
            
            # === Сохранение в JSON ===
            if cards:
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(cards, f, ensure_ascii=False, indent=2)
                print(f"\n💾 JSON сохранён: {json_path.resolve()}")
                
                # === Конвертация в Excel ===
                json_to_xlsx(str(json_path), str(xlsx_path))
            else:
                print("⚠ Нет карточек для сохранения")
            
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            import traceback
            traceback.print_exc()
        finally:
            context.close()
            browser.close()
    
    return cards


if __name__ == "__main__":
    # === НАСТРОЙКИ ===
    TARGET_URL = "https://www.vtb.ru/personal/platezhi-i-perevody"
    HEADLESS = False
    WAIT_TIME = 3.0
    DETAIL_WAIT_TIME = 2.5
    DELAY_BETWEEN_REQUESTS = 2.0
    OUTPUT_DIR = None  # None = ./vtb_downloads/ рядом со скриптом
    
    # === ЗАПУСК ===
    result = main(
        url=TARGET_URL,
        headless=HEADLESS,
        output_dir=OUTPUT_DIR,
        wait_time=WAIT_TIME,
        detail_wait_time=DETAIL_WAIT_TIME,
        delay_between_requests=DELAY_BETWEEN_REQUESTS
    )
    
    # === Вывод результата ===
    if result:
        print(f"\n📋 Итоговый результат содержит {len(result)} карточек")
        for card in result[:2]:
            print(f"\n{'='*60}")
            print(f"📌 {card['description']}")
            print(f"🏷  {card['website_name'][:60]}...")
            print(f"🔗 {card['Link']}")
            print(f"   📦 elements_list: {len(card['elements_list'])} шт.")
            print(f"   ❓ FAQ: {len(card['FAQ'])} шт.")
    else:
        print("\n⚠ Карточки не найдены")